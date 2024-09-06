import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def extract_cid(identifier):
    # Construct the URL for the detailed phytochemical page
    detailed_page_url = f"https://cb.imsc.res.in/imppat/phytochemical-detailedpage/{identifier}"
    
    # Send the request to the detailed page
    response = requests.get(detailed_page_url)
    
    # Respectful scraping: Add a delay between requests
    time.sleep(2)
    
    # Check if the request was successful
    if response.status_code != 200:
        print(f"Failed to retrieve the CID data. Status code: {response.status_code}")
        return None
    
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Extract the CID from the page
    external_identifiers_section = soup.find(string="External chemical identifiers:")
    if external_identifiers_section:
        cid_tag = external_identifiers_section.find_next('a')
        if cid_tag:
            # Extract the CID from the text, assuming the format is 'CID:XXXX'
            cid_text = cid_tag.text.strip()
            cid_match = re.search(r'CID:(\d+)', cid_text)
            if cid_match:
                cid = cid_match.group(1)
                print(f"Extracted CID: {cid}")
                return cid
    
    print("CID not found")
    return None

def get_smiles_from_pubchem(cid):
    # Construct the URL for the PubChem PUG REST API
    pubchem_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/CanonicalSMILES/TXT"
    print(f"PubChem URL: {pubchem_url}")
    
    # Send the request to the PubChem API
    response = requests.get(pubchem_url)
    
    # Respectful scraping: Add a delay between requests
    time.sleep(2)
    
    # Check if the request was successful
    if response.status_code != 200:
        print(f"Failed to retrieve the SMILES data from PubChem. Status code: {response.status_code}")
        return "N/A"
    
    # Return the SMILES string
    return response.text.strip()

def download_structure(identifier, structure_folder):
    # Construct the URL for the 3D structure
    structure_url = f"https://cb.imsc.res.in/imppat/images/3D/MOL/{identifier}_3D.mol"
    
    # Send the request to download the structure
    response = requests.get(structure_url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Save the structure file
        with open(os.path.join(structure_folder, f"{identifier}_3D.mol"), 'wb') as file:
            file.write(response.content)
        print(f"Downloaded 3D structure for {identifier}")
        return "Downloaded"
    else:
        print(f"Failed to download 3D structure for {identifier}")
        return "Not found"

def search_plant(plant_name, structure_folder):
    # Construct the URL for the specific plant page
    plant_page_url = f"https://cb.imsc.res.in/imppat/phytochemical/{plant_name.replace(' ', '%20')}"
    
    # Send the request to the plant page
    response = requests.get(plant_page_url)
    
    # Respectful scraping: Add a delay between requests
    time.sleep(2)
    
    # Check if the request was successful
    if response.status_code != 200:
        print(f"Failed to retrieve the search results. Status code: {response.status_code}")
        return []
    
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Extract the relevant information from the table
    plant_info = []
    table = soup.find('table')
    if table:
        for row in table.find_all('tr')[1:]:  # Skip the header row
            cols = row.find_all('td')
            if len(cols) > 3:  # Ensure the row has the necessary columns
                identifier = cols[2].text.strip()
                phytochemical_name = cols[3].text.strip()
                print(f"Processing identifier: {identifier}")
                
                # Retrieve CID from the detailed page
                cid = extract_cid(identifier)
                
                # Retrieve SMILES from PubChem
                if cid:
                    smiles = get_smiles_from_pubchem(cid)
                    
                    # Download the 3D structure
                    structure_status = download_structure(identifier, structure_folder)
                    
                    # Prepare the hyperlink if the structure is not found
                    hyperlink = f"https://cb.imsc.res.in/imppat/phytochemical-detailedpage/{identifier}" if structure_status == "Not found" else ""
                    
                    # Append the information to the plant_info list
                    plant_info.append([identifier, phytochemical_name, cid, smiles, structure_status, hyperlink])
    
    return plant_info

def save_to_excel(data, file_path):
    # Add serial number to the data
    data_with_slno = [[i+1] + row for i, row in enumerate(data)]
    
    # Create a DataFrame and save to Excel
    df = pd.DataFrame(data_with_slno, columns=['Sl No', 'Phytochemical Identifier', 'Phytochemical Name', 'CID', 'Canonical SMILES', 'Structure', 'Hyperlink'])
    df.to_excel(file_path, index=False)
    
    # Load the workbook and access the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define fill colors for the structure status cells
    fill_downloaded = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    fill_not_found = PatternFill(start_color="FF3300", end_color="FF3300", fill_type="solid")
    
    # Initialize counters for downloaded and not found compounds
    downloaded_count = 0
    not_found_count = 0
    
    # Apply fill colors based on structure status
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if cell.value == "Downloaded":
                cell.fill = fill_downloaded
                downloaded_count += 1
            else:
                cell.fill = fill_not_found
                not_found_count += 1
    
    # Apply hyperlinks to the hyperlink column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value:
                cell.hyperlink = cell.value
                cell.value = "Link"
                cell.style = "Hyperlink"
    
    # Calculate the total number of compounds
    total_compounds = downloaded_count + not_found_count
    
    # Add summary rows at the end of the sheet
    summary_row = ws.max_row + 2
    ws[f'A{summary_row}'] = "No of compounds downloaded:"
    ws[f'B{summary_row}'] = downloaded_count
    ws[f'A{summary_row+1}'] = "No of missing compounds:"
    ws[f'B{summary_row+1}'] = not_found_count
    ws[f'A{summary_row+2}'] = "Total no of compounds:"
    ws[f'B{summary_row+2}'] = total_compounds
    
    # Save the workbook
    wb.save(file_path)

def main():
    # Prompt the user to enter the save location and validate it
    while True:
        save_location = input("Enter the location to save the file (e.g., D:/path/to/directory/): ")
        if os.path.isdir(save_location):
            break
        else:
            print("Invalid directory. Please enter a valid location.")
    
    # Prompt the user to enter the plant name
    plant_name = input("Enter the name of the Indian medicinal plant: ")
    
    # Ensure the file name is valid
    safe_plant_name = ''.join(c for c in plant_name if c.isalnum() or c in (' ', '_')).rstrip()
    
    # Create a folder named after the plant name in the given location
    plant_folder = os.path.join(save_location, safe_plant_name)
    os.makedirs(plant_folder, exist_ok=True)
    
    # Create a folder named "3d_structure" within the plant folder
    structure_folder = os.path.join(plant_folder, "3d_structure")
    os.makedirs(structure_folder, exist_ok=True)
    
    # Define the full file path for the Excel file
    file_name = f"{safe_plant_name}.xlsx"
    file_path = os.path.join(plant_folder, file_name)
    
    # Search for the plant information
    plant_info = search_plant(plant_name, structure_folder)
    
    # Save the plant information to the Excel file
    if plant_info:
        save_to_excel(plant_info, file_path)
        print(f"Data has been saved to {file_path}")
    else:
        print("No data found for the given plant name.")

if __name__ == "__main__":
    main()
